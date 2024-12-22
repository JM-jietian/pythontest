# -*- coding:utf-8 -*-
# @Time: 2023/5/26  12：00
# @Author: J
import xmind
import os
from openpyxl import load_workbook


class Convert:
    def gen_xmind_file(self, xlsx, output_folder, sheet_name_to_process):
        # 使用 openpyxl 加载 Excel 文件
        readbook = load_workbook(xlsx)
        sheet_list = readbook.sheetnames  # 获取 Excel 中所有的工作表名

        # 遍历每个工作表
        for sheet_name in sheet_list:
            if sheet_name != sheet_name_to_process:
                continue  # 跳过不需要处理的工作表

            # 使用 xmind 加载或创建 XMind 文件
            workbook = xmind.load(sheet_name + '.xmind')
            first_sheet = workbook.getPrimarySheet()
            root_topic = first_sheet.getRootTopic()
            root_topic.setTitle(sheet_name)  # 设置根主题的标题为工作表名
            sheet = readbook[sheet_name]  # 获取当前工作表

            # 遍历工作表的每一行
            for row in sheet.iter_rows(min_row=2):  # 从第二行开始遍历，第一行通常是标题行
                sub_topic = root_topic  # 初始子主题为根主题
                for cell in row:
                    value = cell.value  # 获取单元格的值
                    sub_topic = self.get_or_create_sub_topic(sub_topic, value)

            output_path = os.path.join(output_folder + 'xmind')  # 构建 XMind 文件的输出路径
            xmind.save(workbook, path=output_path)  # 保存 XMind 文件
            print(f"已生成 XMind 文件: {output_path}")

        return 'OK'

    @staticmethod
    def get_or_create_sub_topic(parent_topic, title):
        """
        获取或创建子主题
        :param parent_topic: 父主题
        :param title: 子主题标题
        :return: 子主题对象
        """
        for sub_topic in parent_topic.getSubTopics():
            if sub_topic is not None:
                sub_topic_title = sub_topic.getTitle()
                if sub_topic_title is not None:
                    sub_topic_title = str(sub_topic_title)  # 强制将主题标题转换为字符串
                    if sub_topic_title == title:
                        return sub_topic

        new_sub_topic = parent_topic.addSubTopic()  # 创建新的子主题
        new_sub_topic.setTitle(title if title is not None else '')  # 设置子主题的标题，如果为None则设置为空字符串
        return new_sub_topic


# 在主程序中指定要处理的工作表名称
if __name__ == '__main__':
    global output_path
    # 文件路径
    excel_file = r"C:\Users\86173\Desktop\10.15.0.xlsx"
    xmind_folder = excel_file.split('xlsx')[0]
    sheet_name_to_process = "Sheet1"
    converter = Convert()
    converter.gen_xmind_file(excel_file, xmind_folder, sheet_name_to_process)
