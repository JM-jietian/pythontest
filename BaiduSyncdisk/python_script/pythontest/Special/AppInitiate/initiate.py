# !/usr/bin/python
# -*- coding: utf-8 -*-
# @Time: 2024/12/10
# @Author: J
import os
import re
import subprocess
import threading
import tkinter as tk
from tkinter import filedialog, ttk
from PIL import Image, ImageTk
from openpyxl import Workbook
import io
import sys
import logging
from datetime import datetime


class Method:

    def __init__(self, ):
        self.log_file = 'LOG.log'
        # 初始化日志记录器
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)

        # 创建文件处理器，并设置级别为INFO
        # fh = logging.FileHandler(self.log_file, mode='w', encoding='utf-8')  # 存储到日志文件
        fh = logging.StreamHandler()  # 控制台输出
        fh.setLevel(logging.INFO)

        # 创建并设置日志格式
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)

        # 添加处理器到日志记录器
        self.logger.addHandler(fh)

    @staticmethod
    def connection_status(status_func):
        # 定义装饰器函数，判断设备连接状态
        def device(*args, **kwargs):
            devices = [line.split('\t')[0].strip() for line in os.popen('adb devices').readlines()[1:] if line.split('\t')[0].strip()]
            num_devices = len(devices)
            if num_devices == 0:
                print('暂无设备连接')
            elif num_devices > 1:
                print('当前连接多台设备，仅支持单设备操作，请检查！')
            else:
                status_func(*args, **kwargs)

        return device

    def sub(self, cmd, callback=None):
        # 封装处理subprocess
        self.logger.info(f'执行指令：{cmd}')
        try:
            output = subprocess.check_output(cmd, shell=True, stderr=subprocess.STDOUT).decode('utf-8', errors='ignore')
            self.logger.info(output)
            if callback:
                callback(output)
            return output.strip()
        except subprocess.CalledProcessError as e:
            self.logger.error("Command '%s' failed with error: %s", cmd, e.output.decode('utf-8', errors='ignore'))
            if callback:
                callback(None)

    def async_sub(self, cmd):
        # 异步执行subprocess
        threading.Thread(target=self.sub, args=(cmd,)).start()

    @staticmethod
    def find_file_partial(file_pattern, search_path='/'):
        # 输出指定文件完整路径
        found = None
        found_event = threading.Event()

        def search_files():
            nonlocal found
            for root, dirs, files in os.walk(search_path):
                for file_name in files:
                    if re.search(file_pattern, file_name):
                        found = os.path.abspath(root)  # 返回文件的完整路径
                        found_event.set()  # 通知主线程
                        return

        thread = threading.Thread(target=search_files)
        thread.start()
        found_event.wait()  # 等待搜索完成
        thread.join()  # 确保线程结束
        return found

    @staticmethod
    def write_to_excel(data, file_path):
        # 创建一个工作簿
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        # 写入标题行
        for col, value in enumerate(data[0]):
            ws.cell(row=1, column=col + 1, value=value)
        # 写入数据行
        for row, row_data in enumerate(data[1:], start=2):  # 从第二行开始写入
            for col, value in enumerate(row_data):
                ws.cell(row=row, column=col + 1, value=value)

        # 保存工作簿到文件
        wb.save(file_path)

    def delete_files_with_extension(self, folder_path, extension):
        # 检查文件夹路径是否存在
        if not os.path.exists(folder_path):
            self.logger.error("文件夹路径不存在")
            return

        # 确保扩展名以点开头
        if not extension.startswith('.'):
            extension = '.' + extension

        # 遍历文件夹及其子文件夹中的所有文件
        for root, dirs, files in os.walk(folder_path):
            for filename in files:
                file_path = os.path.join(root, filename)

                # 检查文件后缀名
                if filename.endswith(extension):
                    try:
                        os.unlink(file_path)
                        self.logger.info(f"已删除文件：{file_path}")
                    except Exception as e:
                        self.logger.error(f"删除文件 {file_path} 时出错：{e}")

    @staticmethod
    def get_desktop_path():
        # 获取当前操作系统类型
        operating_system = os.name

        # 根据操作系统获取桌面路径
        if operating_system == 'posix':  # Linux, Unix, macOS
            desktop_path = os.path.expanduser('~/Desktop')
        elif operating_system == 'nt':  # Windows
            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        else:
            # 其他操作系统暂不考虑
            desktop_path = None

        return desktop_path

    def add_to_path_if_not_exists(self, directory_path):
        # 获取系统PATH环境变量
        paths = os.environ.get('PATH', '')
        # 根据操作系统使用正确的路径分隔符
        path_separator = os.pathsep
        # 检查目录路径是否已经在PATH中
        if directory_path in paths.split(path_separator):
            self.logger.info(f"路径 {directory_path} 已在PATH中。")
            return

        # 如果路径不在PATH中，将其添加进去
        if not paths.endswith(path_separator):
            new_path = paths + path_separator + directory_path
        else:
            new_path = paths + directory_path

        # 更新环境变量，临时添加，只会影响当前进程及其子进程
        os.environ['PATH'] = new_path
        self.logger.info(f"路径 {directory_path} 临时添加到PATH环境变量中。")

        # 永久添加需要管理员权限运行Python脚本
        # try:
        #     import ctypes
        #     ctypes.windll.kernel32.SetEnvironmentVariableW("PATH", new_path)
        #     self.logger.info(f"路径 {directory_path} 永久添加到系统PATH环境变量中。")
        # except Exception as e:
        #     self.logger.error(f"添加路径到系统PATH环境变量失败：{e}")


class TextRedirector(io.TextIOBase):
    # 创建自定义输出流类：继承自io.TextIOBase，并重写write方法以将输出定向到Text控件
    def __init__(self, widget):
        self.widget = widget

    def write(self, message):
        self.widget.insert(tk.END, message)
        self.widget.see(tk.END)  # 自动滚动到底部


class ImageBrowser(Method):
    def __init__(self, root):
        # 设置窗口标题和大小
        super().__init__()
        self.root = root
        self.root.title("启动速度测试")
        self.root.geometry("1200x650+100+100")
        # self.root.resizable(False, False)  # 禁止窗口调整大小

        # 创建样式对象，设置主题
        self.style = ttk.Style()
        self.style.theme_use('clam')

        # 创建可分割窗口
        self.paned_window = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill=tk.BOTH, expand=True)

        # 创建图片显示区域
        self.image_frame = ttk.Frame(self.paned_window)
        self.image_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 创建Label控件用于显示图片
        self.label = tk.Label(self.image_frame, text="图片显示区域", font=("Helvetica", 16), compound="center")
        self.label.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 创建文本框和滚动条
        self.text_frame = ttk.Frame(self.paned_window)
        self.text_frame.pack(side=tk.RIGHT, fill=tk.Y)
        self.output_text = tk.Text(self.text_frame, height=10, width=120, font=("Helvetica", 10))
        self.output_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar = tk.Scrollbar(self.text_frame, command=self.output_text.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill='y')
        self.output_text.config(yscrollcommand=self.scrollbar.set)

        # 创建重定向器
        self.redirector = TextRedirector(self.output_text)
        # 重定向sys.stdout
        sys.stdout = self.redirector

        # 创建控件容器
        self.control_frame = ttk.Frame(self.root)
        self.control_frame.pack(side=tk.BOTTOM, fill=tk.X)

        # 初始化其他属性
        self.image_files = []  # 存储图片文件
        self.current_image_index = 0  # 当前图片索引
        self.start_frame = None  # 初始帧
        self.end_frame = None  # 结束帧
        self.screen_recording_time = None  # 录屏时长
        self.input_value = None  # 测试场景
        self.startup_duration = None  # 启动耗时
        self.entry = None  # 文本输入款内容
        self.paths = {}  # 路径

        # 创建控件
        self.create_widgets()

        # 添加窗口大小变化的事件监听器
        self.root.bind('<Configure>', self.on_resize)

    def create_widgets(self):
        # 创建输入框和标签
        label_input = ttk.Label(self.control_frame, text="测试场景:", font=("Helvetica", 12))
        label_input.pack(side=tk.LEFT, padx=5, expand=True)
        self.entry = ttk.Entry(self.control_frame, width=35, font=("Helvetica", 10))
        self.entry.pack(side=tk.LEFT, padx=5, expand=True)

        # 创建按钮
        initiate_button = ttk.Button(self.control_frame, text="开始录屏",
                                     command=lambda: threading.Thread(target=self.initiate).start(), style='TButton')
        initiate_button.pack(side=tk.LEFT, padx=5, expand=True)

        # open_button = ttk.Button(self.control_frame, text="选择初始/结束帧", command=self.open_folder, style='TButton')
        # open_button.pack(side=tk.LEFT, padx=5, expand=True)

        prev_button = ttk.Button(self.control_frame, text="上一张",
                                 command=lambda: threading.Thread(target=self.show_prev_image).start(), style='TButton')
        prev_button.pack(side=tk.LEFT, padx=5, expand=True)

        next_button = ttk.Button(self.control_frame, text="下一张",
                                 command=lambda: threading.Thread(target=self.show_next_image).start(), style='TButton')
        next_button.pack(side=tk.LEFT, padx=5, expand=True)

        start_button = ttk.Button(self.control_frame, text="设置初始帧",
                                  command=lambda: threading.Thread(target=self.set_start_frame).start(),
                                  style='TButton')
        start_button.pack(side=tk.LEFT, padx=5, expand=True)

        end_button = ttk.Button(self.control_frame, text="设置结束帧",
                                command=lambda: threading.Thread(target=self.set_end_frame).start(), style='TButton')
        end_button.pack(side=tk.LEFT, padx=5, expand=True)

        compute_button = ttk.Button(self.control_frame, text="计算耗时",
                                    command=lambda: threading.Thread(target=self.compute).start(), style='TButton')
        compute_button.pack(side=tk.LEFT, padx=5, expand=True)

        save_button = ttk.Button(self.control_frame, text="保存结果",
                                 command=lambda: threading.Thread(target=self.save).start(), style='TButton')
        save_button.pack(side=tk.LEFT, padx=5, expand=True)

    def get_input_value(self):
        # 获取输入框中的值
        return self.entry.get()

    def on_resize(self, event):
        # 获取窗口的当前大小
        width, height = self.root.winfo_width(), self.root.winfo_height()

        # 设置最小窗口大小，这里设置为800x600
        min_width = 1200
        min_height = 650

        # 如果窗口大小小于最小值，则调整为最小值
        if width < min_width:
            self.root.geometry(f'{min_width}x{min_height}+{self.root.winfo_x()}+{self.root.winfo_y()}')
        if height < min_height:
            self.root.geometry(f'{min_width}x{min_height}+{self.root.winfo_x()}+{self.root.winfo_y()}')

        # 确保控件能够适应新的窗口大小
        self.root.update_idletasks()

    @Method.connection_status
    def initiate(self):
        print("正在启动投屏 ······ \n注意：操作完成后关闭投屏即可")
        self.paths.update(
                {
                    "desktop": self.get_desktop_path(),
                    "scrcpy": self.find_file_partial("scrcpy-console.bat"),
                    "ffmpeg": self.find_file_partial("ffmpeg_initiate"),
                    "video": self.find_file_partial("ffmpeg_video"),
                    "image": self.find_file_partial("ffmpeg_Image")
                }
            )
        # 判断 scrcpy，ffmpeg是否添加了环境变量，如果没有则临时添加
        self.add_to_path_if_not_exists(self.paths['scrcpy'])
        self.add_to_path_if_not_exists(self.paths['ffmpeg'])
        # 开始录屏并保存视频
        self.sub(rf"scrcpy --record {self.paths['video']}\video.mp4")
        # 获取录屏时长
        self.screen_recording_time = self.sub(
            rf'ffprobe -v quiet -show_entries format=duration -of default=noprint_wrappers=1:nokey=1 "{self.paths["video"]}\video.mp4"')
        print("录屏时长: %s 秒" % self.screen_recording_time)
        # 拆解录屏，生成每一帧的图片
        self.delete_files_with_extension(self.paths['image'], '.png')  # 拆解前先删除之前的图片文件
        self.sub(rf"ffmpeg -i {self.paths['video']}\video.mp4 -r 10 {self.paths['image']}\%05d.png")
        # 获取拆解后图片总数
        self.image_files = [os.path.join(self.paths['image'], f) for f in os.listdir(self.paths['image'])
                            if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif'))]
        print("拆解图片： %s 张" % len(self.image_files))

        # 拆解结束后展示图片
        self.open_folder()

    def open_folder(self):
        # 打开文件夹选择对话框，并设置默认路径
        # folder_path = filedialog.askdirectory(initialdir=self.paths['image'])
        folder_path = self.paths['image']
        if folder_path:
            # 更新图片文件列表
            self.image_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path)
                                if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif'))]
            # 根据新的图片列表更新初始帧和结束帧
            self.start_frame = 0
            self.end_frame = len(self.image_files) - 1
            # 重置当前图片索引为0
            self.current_image_index = 0
            # 显示第一张图片
            self.show_image(self.image_files[self.current_image_index])

    def show_image(self, image_path):
        # 显示图片，考虑初始帧和结束帧
        if self.start_frame is not None and self.end_frame is not None:  # 如果设置了初始帧和结束帧
            index = self.image_files.index(image_path)  # 获取当前图片的索引
            if index < self.start_frame or index > self.end_frame:  # 如果图片不在初始帧和结束帧之间
                return  # 不显示图片
        img = Image.open(image_path)  # 打开图片文件
        img.thumbnail((800, 600))  # 调整图片大小以适应窗口
        photo = ImageTk.PhotoImage(img)  # 将图片转换为Tkinter可以显示的格式
        self.label.config(image=photo)  # 将图片显示在Label控件上
        self.label.image = photo  # 保持对图片的引用，防止被垃圾回收

    @Method.connection_status
    def show_prev_image(self):
        # 显示上一张图片，考虑初始帧和结束帧
        if self.image_files and self.start_frame is not None and self.end_frame is not None:  # 如果有图片文件并且设置了初始帧和结束帧
            self.current_image_index = (self.current_image_index - 1) % len(self.image_files)  # 计算上一张图片的索引
        self.show_image(self.image_files[self.current_image_index])  # 显示上一张图片

    @Method.connection_status
    def show_next_image(self):
        # 显示下一张图片，考虑初始帧和结束帧
        if self.image_files and self.start_frame is not None and self.end_frame is not None:  # 如果有图片文件并且设置了初始帧和结束帧
            self.current_image_index = (self.current_image_index + 1) % len(self.image_files)  # 计算下一张图片的索引
        self.show_image(self.image_files[self.current_image_index])  # 显示下一张图片

    @Method.connection_status
    def set_start_frame(self):
        # 设置初始帧
        if self.image_files:  # 如果有图片文件
            self.start_frame = self.current_image_index  # 将当前图片索引设置为初始帧
            print("设置初始帧: %s \t 【%s】" % (self.start_frame, self.image_files[self.start_frame]))

    @Method.connection_status
    def set_end_frame(self):
        # 设置结束帧
        if self.image_files:  # 如果有图片文件
            self.end_frame = self.current_image_index  # 将当前图片索引设置为结束帧
            print("设置结束帧: %s \t 【%s】" % (self.end_frame, self.image_files[self.end_frame]))

    @Method.connection_status
    def compute(self):
        # 计算启动耗时
        self.input_value = self.get_input_value()
        self.start_frame += 1
        self.end_frame += 1
        # 如果测试场景为空设置默认值为当前时间
        if self.input_value is '':
            now = datetime.now()
            self.input_value = now.strftime("%Y_%m_%d_%H_%M_%S")
        if self.screen_recording_time and len(self.image_files) and self.start_frame and self.end_frame:
            # 计算时间：视频总时间/拆解视频图片总帧数*间隔帧数
            self.startup_duration = float(self.screen_recording_time) / int(len(self.image_files)) * int(self.end_frame - self.start_frame)
            print("耗时：%s 秒" % self.startup_duration)
        else:
            print("参数异常！\n\n录屏时长: %s \n图片总数: %s \n开始帧: %s \n结束帧: %s \n"
                  % (self.screen_recording_time, len(self.image_files), self.start_frame, self.end_frame))

    @Method.connection_status
    def save(self):
        # 保存测试结果
        self.write_to_excel([
            ('测试场景', '录屏时长', '图片总数', '启动耗时'),
            (self.input_value, self.screen_recording_time, len(self.image_files), self.startup_duration)
        ],
            rf"{self.paths['desktop']}\{self.input_value}.xlsx")
        print(rf"测试结果已保存至： {self.paths['desktop']}\{self.input_value}.xlsx")


if __name__ == "__main__":
    root = tk.Tk()
    app = ImageBrowser(root)  # 创建ImageBrowser类的实例
    root.mainloop()  # 进入Tkinter事件循环