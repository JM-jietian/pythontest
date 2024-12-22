# -*- coding:utf-8 -*-
# @Time: 2023/5/28  10：00
# @Author: TJ
import xmind
from openpyxl import load_workbook
import os
import subprocess
import random
import threading
import time
from tkinter import scrolledtext
import tkinter as tk
from tkinter import ttk
import pandas as pd
import xml.etree.ElementTree as ET
from pathlib import Path

# adb安装路径
adb_path = 'adb'

if "ANDROID_HOME" in os.environ:
    command = os.path.join(
        os.environ["ANDROID_HOME"],
        "platform-tools",
        "adb")
else:
    raise EnvironmentError(
        "在 $ANDROID_HOME 中找不到 Adb : %s." %
        os.environ["ANDROID_HOME"])


class Encapsulation:
    @staticmethod
    def clear_text():
        """清空多行文本框内容"""
        Output.delete(1.0, tk.END)

    @staticmethod
    def current_time():
        """获取当前时间"""
        localtime = time.localtime(time.time())
        localtime = time.strftime("%Y-%m-%d %H:%M:%S", localtime)
        return localtime

    @staticmethod
    def output_insert(txt):
        Output.insert(tk.END, rf"{txt}")
        root.update()  # 更新文本框
        Output.see(tk.END)  # 将文本框滚动到最底部

    @staticmethod
    def invoke(cmd):
        output, errors = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE).communicate()
        oe = output.decode("utf-8")
        return oe

    @staticmethod
    def devices():
        """if 条件语句用于筛选出不为空的设备号。因为如果 split('\t')[0] 之后得到的字符串为空，则说明这一行并不是设备号，而是其它信息（如标题行等），需要被过滤掉"""
        return [line.split('\t')[0].strip() for line in os.popen('adb devices').readlines()[1:] if
                line.split('\t')[0].strip()]

    @staticmethod
    def device_connections():
        # 判断有无设备连接
        num_devices = len(Encapsulation.devices())
        if num_devices == 0:
            return Encapsulation.output_insert(f"{Encapsulation.current_time()}\t没有已连接的设备.\n")

    @staticmethod
    def get_desktop_path():
        # 获取当前操作系统类型
        operating_system = os.name
        # 根据操作系统获取桌面路径
        if operating_system == 'posix':  # Linux, Unix, macOS
            desktop_path = os.path.expanduser('~/Desktop')
        elif operating_system == 'nt':  # Windows
            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        else:
            # 其他操作系统暂不考虑
            desktop_path = None
        return desktop_path

    @staticmethod
    def kill_server():
        os.system(f"adb kill-server & adb start-server")  # 重启adb服务，断开全部连接
        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t已重启adb服务~\n")

    @staticmethod
    def update_combobox(combobox, files):
        # 清空下拉框
        combobox['values'] = ()
        # 更新下拉框的选项
        combobox['values'] = files

    @staticmethod
    def version_number(cmd):
        output, errors = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE).communicate()
        if errors:
            e = errors.decode("gbk")
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t{e}\n")
            return e
        else:
            o = output.decode("gbk")
            return o

    @staticmethod
    def start_threading(Function):
        start_install = threading.Thread(target=Function)
        start_install.start()


class ADB:
    # 参数: device
    def __init__(self, device):
        self.device = "-s %s" % device

    def adb(self, args):
        cmd = "%s %s %s" % (adb_path, self.device, str(args))
        return Encapsulation.invoke(cmd)

    def shell(self, args):
        cmd = "%s %s shell %s" % (adb_path, self.device, str(args))
        return Encapsulation.invoke(cmd)

    def get_device_state(self):
        """
        获取设备状态： offline | bootloader | device
        """
        return self.adb("get-state").strip()

    def get_device(self):
        """
        获取设备id号，return serialNo
        """
        return self.adb("get-serialno").strip()

    def get_android_version(self, is_big_version=False):
        """
        获取设备中的Android版本号，如4.2.2
        如果is_big_version=True 则只获取大版本号，如1-14
        """
        return self.shell("getprop ro.build.version.release").strip()

    def get_sdk_version(self):
        """
        获取设备SDK版本号，如：24
        """
        return self.shell("getprop ro.build.version.sdk").strip()

    def get_product_brand(self):
        """
        获取设备品牌，如：HUAWEI
        """
        return self.shell("getprop ro.product.brand").strip()

    def get_product_model(self):
        """
        获取设备型号，如：MHA-AL00
        """
        return self.shell("getprop ro.product.model").strip()

    def get_product_rom(self):
        """
        获取设备ROM名，如：MHA-AL00C00B213
        """
        return self.shell("getprop ro.build.display.id").strip()


class Detect:
    """检测设备连接状态"""

    @staticmethod
    def detect():
        Encapsulation.device_connections()  # 判断有无设备连接
        N = 1
        for device in Encapsulation.devices():
            adb_d = ADB(device)
            if not adb_d.get_device_state():
                Encapsulation.output_insert(f'{Encapsulation.current_time()}\t设备{device}\t已断开连接\n')
                continue
            else:
                Encapsulation.output_insert(f"{Encapsulation.current_time()}\t序列:{N}\n"
                                            f"{Encapsulation.current_time()}\t设备ID:{adb_d.get_device()}\n"
                                            f"{Encapsulation.current_time()}\t设备连接状态:{adb_d.get_device_state()}\n"
                                            f"{Encapsulation.current_time()}\tAndroid版本号:{adb_d.get_android_version()}\n"
                                            f"{Encapsulation.current_time()}\tSDK版本号:{adb_d.get_sdk_version()}\n"
                                            f"{Encapsulation.current_time()}\t设备品牌:{adb_d.get_product_brand()}\n"
                                            f"{Encapsulation.current_time()}\t设备型号:{adb_d.get_product_model()}\n"
                                            f"{Encapsulation.current_time()}\t设备ROM名:{adb_d.get_product_rom()}\n")
                N += 1

    @staticmethod
    def detect_threading():
        Encapsulation.start_threading(Detect.detect)


class Process_monitoring:
    """进程监控
    adb shell am force-stop com.tencent.mobileqq 根据包名杀掉指定进程
    """

    def __init__(self):
        self.waiting_message_shown = False
        self.stop_event = threading.Event()  # 用于通知线程停止的 Event 对象
        self.running_process = None  # 用于存储运行的线程对象

    @staticmethod
    def check_device_connection():
        try:
            # 使用subprocess模块执行adb devices命令并获取输出
            cmd = "adb devices"
            result = subprocess.run(cmd, shell=True, capture_output=True, text=True, check=False)
            if result.returncode == 0:
                device_lines = result.stdout.splitlines()
                if len(device_lines) >= 3:
                    return True
        except Exception as e:
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t检查设备连接状态发生错误: {e}")
        return False

    # 写入状态信息到文件
    @staticmethod
    def write_status_to_file(output_file, status):
        CurrentTime = time.strftime("%H:%M:%S", time.localtime())
        with open(output_file, "a") as f:
            f.write(f"{CurrentTime}\t{status}\n")

    # 监控指定APP的进程信息
    def monitoring_app_process(self, package_name):
        desktop = Encapsulation.get_desktop_path()
        output_file = rf"{desktop}\{package_name}_ps.txt"
        with open(output_file, "w") as f:
            f.write("时间戳\t进程信息\n")

        connected = False
        last_disconnected_time = None

        while not self.stop_event.is_set():  # 检查 Event 是否被设置
            try:
                if not connected:
                    if self.check_device_connection():
                        status = "设备已连接"
                    else:
                        status = "等待连接设备..."
                        if not self.waiting_message_shown:
                            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t" + status + "\n")
                            self.waiting_message_shown = True
                        time.sleep(5)
                        continue

                    Encapsulation.output_insert(f"{Encapsulation.current_time()}\t" + status + "\n")
                    connected = True
                    self.waiting_message_shown = False

                if connected:
                    cmd = "adb shell ps | findstr {package_name}".format(package_name=package_name)
                    result = subprocess.run(cmd, shell=True, capture_output=True, text=True, check=False)

                    if result.returncode == 0:
                        process_lines = result.stdout.splitlines()
                    else:
                        process_lines = []

                    CurrentTime = time.strftime("%H:%M:%S", time.localtime())
                    with open(output_file, "a") as f:
                        if not process_lines:
                            f.write(f"{CurrentTime}\t未找到进程\n")
                            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t未找到进程\n")

                        else:
                            for process_info in process_lines:
                                f.write(f"{CurrentTime}\t{process_info}\n")
                                Encapsulation.output_insert(
                                    f"{Encapsulation.current_time()}\t{process_info}\n")

                if connected and not self.check_device_connection():
                    if last_disconnected_time is None or time.time() - last_disconnected_time >= 10:
                        status = "设备已断开连接"
                        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t" + status + "\n")
                        last_disconnected_time = time.time()
                        connected = False
                        self.waiting_message_shown = False
            except Exception as e:
                Encapsulation.output_insert(f"{Encapsulation.current_time()}\t发生错误: {e}")

            time.sleep(3)

    def start_monitoring(self):
        user_input = entry_var.get()
        if user_input is '':
            return Encapsulation.output_insert(f"{Encapsulation.current_time()}\t进程名为必填项！！！\n")
        label.config(text="接收到的输入： " + user_input)
        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t正在运行......\n")
        outcome = "运行结果：\t请勿连接多台设备，仅支持单台设备监控！！！\n"
        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t" + outcome)

        self.stop_event.clear()  # 清除 Event 状态
        self.running_process = threading.Thread(target=self.monitoring_app_process, args=(user_input,))
        self.running_process.start()

    def stop_monitoring(self):
        self.stop_event.set()  # 设置 Event，通知线程停止
        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t监控结束，监控日志已保存至桌面~\n")


class DTC:
    """adb tcpip - connect 连接设备"""

    def __init__(self):
        self.devices_id = None  # 设备devices_id
        self.devices_ip = None  # 设备ip地址
        self.port = None  # 连接端口
        self.ID = None  # 标记
        self.IP = None  # 标记

    def devices_ID(self):
        try:
            # 获取设备devices_id
            self.devices_id = [line.split('\t')[0].strip() for line in os.popen('adb devices').readlines()[1:] if
                               line.split('\t')[0].strip()]
            if not self.devices_id:
                Encapsulation.output_insert(f"{Encapsulation.current_time()}\t没有已连接的设备！！！\n")
                self.ID = True
                return
        except Exception as e:
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t没有已连接的设备！！！\n"
                                        f"{Encapsulation.current_time()}\t{e}\n")
            self.ID = True
            return

    def devices_IP(self, did):
        try:
            # 获取设备ip地址
            self.devices_ip = [line.split('/')[0].strip().strip("inet ") for line in
                               os.popen(f'adb -s {did} shell ip addr show wlan0 | findstr "inet"').readlines()[:1]]
            if not self.devices_id:
                Encapsulation.output_insert(f"{Encapsulation.current_time()}\t没有已连接的设备！！！\n")
                self.IP = True
                return
        except Exception as e:
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t设备没有网络连接！！！\n"
                                        f"{Encapsulation.current_time()}\t{e}")
            self.IP = True
            return

    def tcp_server(self):
        self.devices_ID()  # 获取设备id
        if self.ID:  # 标记为True时结束运行
            return
        try:
            for d_id in self.devices_id:  # 循环创建连接
                Encapsulation.output_insert(f"设备ID：{d_id}\t开始建立连接~\n")
                self.devices_IP(d_id)  # 获取设备ip地址
                if self.IP:  # 标记为True时结束运行
                    return
                self.port = random.randrange(3000, 65500, 2)  # 随机数，作为连接端口用
                # adb tcpip & adb connect创建连接
                cmd = f"adb -s {d_id} tcpip {self.port} & adb -s {d_id} connect {self.devices_ip[0]}:{self.port}"
                output, errors = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE,
                                                  stderr=subprocess.PIPE).communicate()
                time.sleep(3)
                if errors:
                    Encapsulation.output_insert(f"{Encapsulation.current_time()}\t连接失败~\n"
                                                f"Device ID:{self.devices_id[0]}\n错误:{errors.decode('utf-8')}\n")
                else:
                    Encapsulation.output_insert(f"{Encapsulation.current_time()}\t连接成功~\n"
                                                f"{Encapsulation.current_time()}\tID：{d_id}\n"
                                                f"{Encapsulation.current_time()}\tIP：{self.devices_ip[0]}\n"
                                                f"{Encapsulation.current_time()}\t接口：{self.port}\n")
        except Exception as e:
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t异常 ！！！ \n"
                                        f"{Encapsulation.current_time()}\t{e}\n")
        except UnicodeDecodeError as e:
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t异常 ！！！ \n"
                                        f"{Encapsulation.current_time()}\t{e}\n")

    def tcp_server_threading(self):
        Encapsulation.start_threading(self.tcp_server)

    @staticmethod
    def kill_server_threading():
        Encapsulation.start_threading(Encapsulation.kill_server)


class DP:
    """dp跳转"""

    @staticmethod
    def dp():
        dp = DpInput.get()
        if not dp:
            return Encapsulation.output_insert(f"{Encapsulation.current_time()}\tDP为必填项！！！\n")
        for device in Encapsulation.devices():
            cmd = f"adb -s {device} shell am start -d {dp}"
            Encapsulation.output_insert(
                f"{Encapsulation.current_time()}\t设备ID：{device}\t执行DP跳转\n")
            Encapsulation.invoke(cmd)


class Guidance:

    @staticmethod
    def boot_guidance(device):
        cmd = (f"adb -s {device} shell settings put global device_provisioned 0 && "
               f"adb -s {device} shell settings put secure user_setup_complete 0 && "
               f"adb -s {device} shell settings get global device_provisioned && "
               f"adb -s {device} shell settings get secure user_setup_complete && "
               f"adb -s {device} shell pm enable com.coloros.bootreg && "
               f"adb -s {device} shell pm enable com.coloros.bootreg/.activity.WelcomePage && "
               "ping -n 5 127.1>nul && "
               f"adb -s {device} shell am start com.coloros.bootreg/.activity.WelcomePage")
        result = Encapsulation.version_number(cmd)
        Encapsulation.output_insert(result)

    @staticmethod
    def run():
        for dev in Encapsulation.devices():
            Encapsulation.start_threading(Guidance.boot_guidance(dev))

    @staticmethod
    def run_threading():
        Encapsulation.start_threading(Guidance.run)


class Padding:
    """存储&内存填充"""

    @staticmethod
    def storage():
        try:
            Encapsulation.device_connections()  # 判断有无设备连接
            for device in Encapsulation.devices():
                adb_d = ADB(device)
                if not adb_d.get_device_state():
                    Encapsulation.output_insert(f'{Encapsulation.current_time()}\t设备{device}\t已断开连接\n')
                    continue
            if len(Encapsulation.devices()) > 1:
                return Encapsulation.output_insert(
                    f"{Encapsulation.current_time()}\t多设备连接，运行此功能时请确保仅连接一台设备！！！\n")
            elif var_FilePath.get().split(';')[0] == '':
                Encapsulation.output_insert(f"{Encapsulation.current_time()}\t默认路径:[/sdcard/TempFolder/]\n")
            try:
                FilePath = '/sdcard/TempFolder/' if var_FilePath.get().split(';')[0] is '' else \
                    var_FilePath.get().split(';')[0]
                FileSize = int(var_FilePath.get().split(';')[1])
                NumberOfFiles = int(var_FilePath.get().split(';')[2])
            except ValueError:
                return Encapsulation.output_insert(
                    f"{Encapsulation.current_time()}\tFileSize & NumberOfFiles 请输入正整数！！！\n")

            storage = 'adb shell df -h /data/media'
            you_storage = subprocess.Popen(storage, text=True, shell=True, stdout=subprocess.PIPE,
                                           stderr=subprocess.PIPE)
            Encapsulation.output_insert(
                f"{Encapsulation.current_time()}\t设备剩余存储：\n{you_storage.communicate()[0]}\n")

            FP = f'adb shell mkdir {FilePath}'
            FP_text = subprocess.Popen(FP, text=True, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            if 'Read-only file system' in FP_text.communicate()[1]:
                return Encapsulation.output_insert(f"{Encapsulation.current_time()}\t"
                                                   f"文件[{FilePath}]\t创建失败，请检查输入的路径是否正确！！！\n")

            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t文件[{FilePath}]\t正在创建中，请稍后...\n")
            Name = 1
            for i in range(NumberOfFiles):
                create = f'adb shell dd if=/dev/zero of={FilePath}FileName{Name} bs=1048576 count={FileSize}'
                create_num = subprocess.Popen(create, text=True, shell=True, stdout=subprocess.PIPE,
                                              stderr=subprocess.PIPE)
                Encapsulation.output_insert(f"{Encapsulation.current_time()}\t{create_num.communicate()[1]}\n")
                Name += 1
            you_storage = subprocess.Popen(storage, text=True, shell=True, stdout=subprocess.PIPE,
                                           stderr=subprocess.PIPE)
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t"
                                        f"创建完成，设备剩余存储：\n{you_storage.communicate()[0]}")
        except Exception as e:
            return Encapsulation.output_insert(f"{Encapsulation.current_time()}\t{e}\t缺少必填参数！！！\n"
                                               f"{Encapsulation.current_time()}\t正确填入格式：文件路径;文件大小M;文件数量\n"
                                               f"{Encapsulation.current_time()}\t文件路径可不填，正确填入格式：;文件大小M;文件数量，文件路径默认[/sdcard/TempFolder/]\n")

    @staticmethod
    def del_storage():
        FilePath = '/sdcard/TempFolder' if var_FilePath.get().split(';')[0] is '' else var_FilePath.get().split(';')[0]
        del_FilePath = f'adb shell rm -r {FilePath}'
        subprocess.Popen(del_FilePath, shell=True)
        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t文件[{FilePath}]\t已清除~\n")

    @staticmethod
    def storage_threading():
        Encapsulation.start_threading(Padding.storage)

    @staticmethod
    def memory():
        try:
            Encapsulation.device_connections()  # 判断有无设备连接
            if len(Encapsulation.devices()) > 1:
                return Encapsulation.output_insert(f"{Encapsulation.current_time()}\t"
                                                   f"多设备连接，运行此功能时请确保仅连接一台设备！！！\n")
            # 获取设备内存信息
            cmd = f'adb shell "cat /proc/meminfo | grep MemAvailable"'
            you_memory = subprocess.Popen(cmd, text=True, shell=True, stdout=subprocess.PIPE,
                                          stderr=subprocess.PIPE)
            memory_num = int(you_memory.communicate()[0].split()[1]) / 1024
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t设备剩余内存: {memory_num}M\n")

            if var_rom_num.get() == '':
                return Encapsulation.output_insert(f"{Encapsulation.current_time()}\t内存占用大小为必填项为！！！\n")
            # 计算要占用的内存大小
            input_rom = int(var_rom_num.get())
            rom_num = input_rom * 1048576
            Encapsulation.output_insert(
                f"{Encapsulation.current_time()}\t设备运行内存占用中，占用大小: {input_rom}M，结束运行后恢复...\n"
                f"【注意：在结束占用前，每次点击开始占用都会叠加占用的内存！！！】\n")

            # 使用 dd 命令占用内存
            rom_cmd = f'adb shell dd if=/dev/zero of=/dev/null bs={rom_num}'
            subprocess.run(rom_cmd, shell=True)
        except Exception as e:
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t发生异常：{str(e)}\n")

    @staticmethod
    def startMemory_threading():
        Encapsulation.start_threading(Padding.memory)

    @staticmethod
    def stopMemory_threading():
        Encapsulation.start_threading(Encapsulation.kill_server)
        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t内存占用结束~\n")


class Convert:
    """excel转xmind"""

    @staticmethod
    def gen_xmind_file(xlsx, output_folder, sheet_name_to_process):
        # 使用 openpyxl 加载 Excel 文件
        readbook = load_workbook(xlsx)
        sheet_list = readbook.sheetnames  # 获取 Excel 中所有的工作表名

        # 遍历每个工作表
        for sheet_name in sheet_list:
            if sheet_name != sheet_name_to_process:
                continue  # 跳过不需要处理的工作表

            # 使用 xmind 加载或创建 XMind 文件
            workbook = xmind.load(sheet_name + '.xmind')
            first_sheet = workbook.getPrimarySheet()
            root_topic = first_sheet.getRootTopic()
            root_topic.setTitle(sheet_name)  # 设置根主题的标题为工作表名
            sheet = readbook[sheet_name]  # 获取当前工作表

            # 遍历工作表的每一行
            for row in sheet.iter_rows(min_row=2):  # 从第二行开始遍历，第一行通常是标题行
                sub_topic = root_topic  # 初始子主题为根主题
                for cell in row:
                    value = cell.value  # 获取单元格的值
                    sub_topic = Convert.get_or_create_sub_topic(sub_topic, value)

            output_path = os.path.join(output_folder + 'xmind')  # 构建 XMind 文件的输出路径
            xmind.save(workbook, path=output_path)  # 保存 XMind 文件
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t已生成 XMind 文件: {output_path}\n")

        return 'OK'

    @staticmethod
    def get_or_create_sub_topic(parent_topic, title):
        """
        获取或创建子主题
        :param parent_topic: 父主题
        :param title: 子主题标题
        :return: 子主题对象
        """
        for sub_topic in parent_topic.getSubTopics():
            if sub_topic is not None:
                sub_topic_title = sub_topic.getTitle()
                if sub_topic_title is not None:
                    sub_topic_title = str(sub_topic_title)  # 强制将主题标题转换为字符串
                    if sub_topic_title == title:
                        return sub_topic

        new_sub_topic = parent_topic.addSubTopic()  # 创建新的子主题
        new_sub_topic.setTitle(title if title is not None else '')  # 设置子主题的标题，如果为None则设置为空字符串
        return new_sub_topic

    @staticmethod
    def run():
        try:
            excel_file = var_excel_file.get()  # Excel 文件路径
            xmind_folder = excel_file.split('xlsx')[0]  # XMind文件夹路径与.xlsx文件在同一个目录下
            sheet_name_to_process = "Sheet1"  # 指定要处理的工作表名称
            Convert.gen_xmind_file(excel_file, xmind_folder, sheet_name_to_process)
        except Exception as e:
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t{e}转换异常，请输入正确的Excel文件路径！！！\n")

    @staticmethod
    def convert_threading():
        Encapsulation.start_threading(Convert.run)


class Freemind_Convert:
    """xmind_excel"""

    @staticmethod
    def parse_node(node):
        node_text = node.attrib.get('TEXT', '')  # 获取节点文本
        children_data = []
        for child in node:
            children_data.extend(Freemind_Convert.parse_node(child))  # 递归解析子节点
        if children_data:
            # 将节点文本和子节点数据合并到一个列表
            return [[node_text] + row for row in children_data]
        else:
            # 只有节点文本
            return [[node_text]]

    @staticmethod
    def free_mind_to_excel(input_file, output_file):
        # 解析FreeMind文件
        tree = ET.parse(input_file)
        roots = tree.getroot()
        # 解析节点数据
        data = Freemind_Convert.parse_node(roots)
        # 确定最大列数
        max_columns = max(len(row) for row in data)
        # 填充缺失的列
        for row in data:
            row.extend([''] * (max_columns - len(row)))
        # 创建DataFrame对象
        df = pd.DataFrame(data)
        # 设置列名
        column_names = ['all', '版本', '编写人员', '需求', '测试点', '一级模块', '二级模块', '前置条件', '操作步骤',
                        '预期结果', '用例严重程度']
        if max_columns < len(column_names):
            column_names = column_names[:max_columns]
        elif max_columns > len(column_names):
            column_names.extend(['all'] * (max_columns - len(column_names)))
        df.columns = column_names

        # 删除列名为all的空列表
        df.drop(['all'], axis=1, inplace=True)

        # 保存为Excel文件
        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t正在保存Excel文件~\n")
        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, encoding='utf-8')
        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t已完成,文件存放路径：{output_file}\n")

    @staticmethod
    def run():
        try:
            free_mind_path = var_excel_file.get()  # mm文件路径
            Freemind_Convert.free_mind_to_excel(free_mind_path,
                                                free_mind_path.split('mm')[0] + 'xlsx')  # 将xlsx文件保存在.mm文件同目录下
        except Exception as e:
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t{e}转换异常，请输入正确的Xmind文件路径！！！\n")

    @staticmethod
    def freemind_convert_threading():
        Encapsulation.start_threading(Freemind_Convert.run)


class Pull_log:
    """日志提取"""

    @staticmethod
    def apkPack():
        if vra_Pull_log.get() is '':
            return Encapsulation.output_insert(f"{Encapsulation.current_time()}\t请输入正确的包名\n")

    @staticmethod
    def pull_music_log():
        """导出WS日志文件保存至首页，并使用Notepad++打开"""
        desktop = Encapsulation.get_desktop_path()
        Encapsulation.device_connections()
        Pull_log.apkPack()
        apkName = vra_Pull_log.get()
        for dev in Encapsulation.devices():
            cmd_pull = (f"adb -s {dev} pull "
                        rf"/storage/emulated/0/Android/data/{apkName}/files/ws/logs/{apkName.split('.')[0]}/{apkName.split('.')[1]}/{apkName.split('.')[2]}/ "
                        rf"{desktop}\{dev} "
                        rf"&& start Notepad++ {desktop}\{dev}\{apkName}.*")
            Installation = Encapsulation.version_number(cmd_pull)
            Encapsulation.output_insert(Installation)

    @staticmethod
    def crash_music_log():
        """抓取crash日志"""
        Encapsulation.device_connections()
        for dev in Encapsulation.devices():
            crash_cmd = rf"chcp 65001 && start adb -s {dev} logcat -b crash"
            print(crash_cmd)
            Installation = Encapsulation.version_number(crash_cmd)
            Encapsulation.output_insert(Installation)

    @staticmethod
    def crash_log_threading():
        Encapsulation.start_threading(Pull_log.crash_music_log)

    @staticmethod
    def system_logs():
        """导出logkit系统日志至桌面"""
        Encapsulation.device_connections()
        Pull_log.apkPack()
        desktop = Encapsulation.get_desktop_path()
        for dev in Encapsulation.devices():
            cmd_pull = rf"chcp 65001 && start adb -s {dev} pull /storage/emulated/0/Android/data/com.oplus.logkit/files/Log/ {desktop}\Log"
            Installation = Encapsulation.version_number(cmd_pull)
            Encapsulation.output_insert(Installation)

    @staticmethod
    def system_log_threading():
        Encapsulation.start_threading(Pull_log.system_logs)


class Install:
    def __init__(self):
        self.folder_path = Path(r"E:/feishu")

    @staticmethod
    def walk_directory(folder_path):
        # 筛选指定路径下的apk文件
        # 使用 rglob 获取所有文件，然后筛选出文件
        files = [file for file in folder_path.rglob('*.apk') if file.is_file()]
        # 返回文件的绝对路径，用于在下拉框中显示
        return [str(file) for file in files]

    @staticmethod
    def on_combobox_select(event):
        # 获取当前选中的文件路径
        selected_file = combobox.get()
        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t当前选中的Apk: {selected_file}\n")

    def updateList(self):
        # 刷新列表
        files = Install.walk_directory(self.folder_path)
        if len(files) == 0:
            Encapsulation.output_insert(f"{Encapsulation.current_time()}\t路径：{self.folder_path} 下无Apk文件\n")
        Encapsulation.update_combobox(combobox, files)

    @staticmethod
    def install():
        if combobox.get() is '':
            return Encapsulation.output_insert(f"{Encapsulation.current_time()}\t无选中Apk文件~\n")
        Encapsulation.device_connections()
        for device in Encapsulation.devices():
            adb_d = ADB(device)
            version = adb_d.get_android_version()  # 安卓版本号
            apk_path = combobox.get()  # 当前选中的apk绝对路径
            Encapsulation.output_insert(
                f"{Encapsulation.current_time()}\t设备ID: %s\n"
                f"{Encapsulation.current_time()}\t安卓版本：%s\n"
                f"{Encapsulation.current_time()}\tApk：%s\n" % (device, version, apk_path))
            apk_name = [name for name in apk_path.split('\\') if ".apk" in name][0]  # 当前选中的apk
            try:
                if '40.' in apk_name:
                    if int(version) >= 11:
                        cmd_install = f'adb -s {device} install -r -d -t {apk_path}'
                        Installation = Encapsulation.version_number(cmd_install)
                        Encapsulation.output_insert(f"{Encapsulation.current_time()}\t{Installation}")
                    else:
                        Encapsulation.output_insert(f"{Encapsulation.current_time()}\tInstallation failed......\n")
                        continue
                elif '50.' in apk_name:
                    if 10 <= int(version) < 11 and 'oppo' in apk_name:
                        pass
                    elif 9 <= int(version) < 10 and 'media' in apk_name and 'old' not in apk_name:
                        pass
                    elif int(version) < 9 and 'mediaold' in apk_name:
                        pass
                    else:
                        Encapsulation.output_insert(f"{Encapsulation.current_time()}\tInstallation failed......\n")
                        continue
                    cmd_install = f'adb -s {device} install -r -d -t {apk_path}'
                    Installation = Encapsulation.version_number(cmd_install)
                    Encapsulation.output_insert(f"{Encapsulation.current_time()}\t{Installation}")
                else:
                    cmd_install = f'adb -s {device} install -r -d -t {apk_path}'
                    Installation = Encapsulation.version_number(cmd_install)
                    Encapsulation.output_insert(f"{Encapsulation.current_time()}\t{Installation}")
            except Exception as e:
                Encapsulation.output_insert(f"{Encapsulation.current_time()}\tInstallation failed......\n"
                                            f"{Encapsulation.current_time()}\t{e}")
                continue

    @staticmethod
    def install_threading():
        Encapsulation.start_threading(Install.install)


if __name__ == "__main__":
    # 创建主窗口
    root = tk.Tk()
    root.title("Assistant")
    # 设置主窗口的最小宽度和最小高度
    root.minsize(width=1000, height=500)
    # 设置主窗口的最大宽度和最大高度
    root.maxsize(width=3000, height=1800)
    # 创建三维样式窗口,(root, relief='sunken', bd=2)
    Output_frame = tk.Frame(root)
    TCP_frame = tk.Frame(root)
    DP_frame = tk.Frame(root)
    Padding_frame = tk.Frame(root)
    rom_frame = tk.Frame(root)
    Process_frame = tk.Frame(root)
    Convert_frame = tk.Frame(root)
    Freemind_Convert_frame = tk.Frame(root)
    Pull_log_frame = tk.Frame(root)
    combobox_frame = tk.Frame(root)

    # 创建类对象，用于执行脚本程序
    start_run = DTC()
    Padding_run = Padding()
    DP_run = DP()
    Process_run = Process_monitoring()
    detect_run = Detect()
    Convert_run = Convert()
    Freemind_Convert_run = Freemind_Convert()
    Pull_log_run = Pull_log()
    install_run = Install()
    guidance_run = Guidance()

    # 创建 StringVar 对象，用于关联 Entry 的文本
    var_FilePath = tk.StringVar()
    var_rom_num = tk.StringVar()
    DpInput = tk.StringVar()
    entry_var = tk.StringVar()
    var_excel_file = tk.StringVar()
    vra_Pull_log = tk.StringVar()

    # 创建一个多行文本框，指定宽度和高度，并允许随着主窗口变化而变化
    Output = scrolledtext.ScrolledText(Output_frame, wrap=tk.WORD, width=60, height=20, font=('微软雅黑', 10), bd=2)

    # 创建DP标签、文本框、按钮
    DP_input = tk.Entry(DP_frame, textvariable=DpInput, bd=2, width=35)
    DP_button = tk.Button(DP_frame, text="DP", command=DP_run.dp, padx=3, font=('微软雅黑', 10), bd=2)

    # 创建 "存储填充" 标签、文本框、按钮
    FilePath_label = tk.Label(Padding_frame,
                              text="存储填充/* 文件存放路径;大小;数量", padx=3, font=('微软雅黑', 10), bd=2,
                              fg='red')
    FilePath_entry = tk.Entry(Padding_frame, textvariable=var_FilePath, bd=2, width=50)
    storage_Button = tk.Button(Padding_frame, text="填充",
                               command=Padding_run.storage_threading, padx=3, font=('微软雅黑', 10), bd=2)
    del_storage_Button = tk.Button(Padding_frame, text="清除",
                                   command=Padding_run.del_storage, padx=3, font=('微软雅黑', 10), bd=2)
    # 创建“内存占用”标签、文本框、按钮
    rom_label = tk.Label(rom_frame, text="内存占用/* 大小(M)", padx=3, font=('微软雅黑', 10), bd=2, fg='red')
    rom_entry = tk.Entry(rom_frame, textvariable=var_rom_num, bd=2, width=50)
    start_rom_Button = tk.Button(rom_frame, text="开始",
                                 command=Padding_run.startMemory_threading, padx=3, font=('微软雅黑', 10), bd=2)
    stop_rom_Button = tk.Button(rom_frame, text="结束",
                                command=Padding_run.stopMemory_threading, padx=3, font=('微软雅黑', 10), bd=2)

    # 创建进程标签、文本框和按钮
    label = tk.Label(Process_frame, text="进程监控/* 进程名", padx=3, font=('微软雅黑', 10), bd=2, fg='red')
    entry = tk.Entry(Process_frame, textvariable=entry_var, width=50, bd=2)
    StartButton = tk.Button(Process_frame, text="开始",
                            command=Process_run.start_monitoring, padx=3, font=('微软雅黑', 10), bd=2)
    StopButton = tk.Button(Process_frame, text="结束",
                           command=Process_run.stop_monitoring, padx=3, font=('微软雅黑', 10), bd=2)

    # 创建 "tcp连接" "连接状态" "重启设备" "重启adb服务" 按钮
    clear_button = tk.Button(TCP_frame, text="清空文本框", command=Encapsulation.clear_text,
                             padx=3, pady=5, font=('微软雅黑', 10), bd=2)
    tcp_detect = tk.Button(TCP_frame, text="建立连接", command=start_run.tcp_server_threading,
                           padx=3, pady=5, font=('微软雅黑', 10), bd=2)
    adb_detect = tk.Button(TCP_frame,
                           text="重启adb", command=start_run.kill_server_threading,
                           padx=3, pady=5, font=('微软雅黑', 10), bd=2)
    start_detect = tk.Button(TCP_frame,
                             text="设备信息", command=detect_run.detect_threading, padx=3, pady=5, font=('微软雅黑', 10),
                             bd=2)
    guidance_detect = tk.Button(TCP_frame,
                                text="开机向导", command=guidance_run.run_threading, padx=3, pady=5,
                                font=('微软雅黑', 10), bd=2)

    # 创建excel_xmind按钮、xmind_excel按钮，文本标题
    xmind_title = tk.Label(Convert_frame, text="Xmind-Excel转换/* 文件地址", padx=3, font=('微软雅黑', 10), bd=2, fg='red')
    xmind_entry = tk.Entry(Convert_frame, textvariable=var_excel_file, bd=2, width=50)
    excel_xmind_button = tk.Button(Convert_frame, text="Excel->Xmind", command=Convert_run.convert_threading, padx=3, pady=5,
                                   font=('微软雅黑', 10), bd=2)
    Freemind_Convert_button = tk.Button(Convert_frame, text="Xmind->Excel",
                                        command=Freemind_Convert_run.freemind_convert_threading,
                                        padx=3, pady=5, font=('微软雅黑', 10), bd=2)

    # App日志提取
    Pull_log_title = tk.Label(Pull_log_frame, text="日志/* app包名", padx=3, font=('微软雅黑', 10), bd=2, fg='red')
    Pull_log_entry = tk.Entry(Pull_log_frame, textvariable=vra_Pull_log, bd=2, width=50)
    Pull_log_button1 = tk.Button(Pull_log_frame, text="WS", command=Pull_log_run.pull_music_log, padx=3,
                                 pady=5,
                                 font=('微软雅黑', 10), bd=2)
    Pull_log_button2 = tk.Button(Pull_log_frame, text="crash", command=Pull_log_run.crash_log_threading, padx=3,
                                 pady=5, font=('微软雅黑', 10), bd=2)
    Pull_log_button3 = tk.Button(Pull_log_frame, text="系统", command=Pull_log_run.system_log_threading, padx=3,
                                 pady=5,
                                 font=('微软雅黑', 10), bd=2)

    # install
    # 创建一个 Combobox 组件
    combobox = ttk.Combobox(Output_frame, font=('微软雅黑', 10))
    # 创建一个刷新按钮
    update_list_Button = tk.Button(Output_frame, text="刷新", command=install_run.updateList,
                                   padx=3, font=('微软雅黑', 10), bd=2)
    # 创建安装按钮
    install_Button = tk.Button(Output_frame, text="安装APK", command=install_run.install_threading,
                               padx=3, font=('微软雅黑', 10), bd=2)
    # 绑定选中事件
    combobox.bind("<<ComboboxSelected>>", install_run.on_combobox_select)

    # 布局管理
    Output.pack(padx=10, pady=10, expand=True, fill="both")  # 多行文本框
    Output_frame.pack(side='left', expand=True, fill="both")
    # install
    combobox.pack(side='left', expand=True, fill="x", padx=(10, 0))
    update_list_Button.pack(side='left', fill="x", padx=(0, 0))
    install_Button.pack(side='left', fill="x", padx=(0, 30))
    combobox_frame.pack(expand=True, fill="both", padx=5)

    # 清空文框、安装apk、tcp连接，adb服务重启，设备重启，设备信息
    clear_button.pack(side='left', expand=True, fill="x")
    tcp_detect.pack(side='left', expand=True, fill="x")
    adb_detect.pack(side='left', expand=True, fill="x")
    start_detect.pack(side='left', expand=True, fill="x")
    guidance_detect.pack(side='left', expand=True, fill="x")
    TCP_frame.pack(expand=True, fill="both", pady=5)

    # DP
    DP_input.pack(side='left', expand=True, fill="x")
    DP_button.pack(side='left', expand=True, fill="x")
    DP_frame.pack(expand=True, fill="both", pady=5)

    # Padding
    FilePath_label.pack()
    FilePath_entry.pack()
    storage_Button.pack(side='left', expand=True, padx=(50, 0), fill="x")
    del_storage_Button.pack(side='left', expand=True, padx=(0, 50), fill="x")
    Padding_frame.pack(expand=True, fill="both", pady=5)

    # memory
    rom_label.pack()
    rom_entry.pack()
    start_rom_Button.pack(side='left', expand=True, padx=(50, 0), fill="x")
    stop_rom_Button.pack(side='left', expand=True, padx=(0, 50), fill="x")
    rom_frame.pack(expand=True, fill="both", pady=5)

    # 进程监控
    label.pack()
    entry.pack()
    StartButton.pack(side='left', expand=True, padx=(50, 0), fill="x")
    StopButton.pack(side='left', expand=True, padx=(0, 50), fill="x")
    Process_frame.pack(expand=True, fill="both", pady=5)

    # excel_xmind、xmind_excel
    xmind_title.pack()
    xmind_entry.pack()
    excel_xmind_button.pack(side='left', expand=True, padx=(50, 0), fill="x")
    Freemind_Convert_button.pack(side='left', expand=True, padx=(0, 50), fill="x")
    Convert_frame.pack(expand=True, fill="both", pady=5)

    # App日志提取
    Pull_log_title.pack()
    Pull_log_entry.pack()
    Pull_log_button1.pack(side='left', expand=True, padx=(0, 0), fill="x")
    Pull_log_button2.pack(side='left', expand=True, padx=(0, 0), fill="x")
    Pull_log_button3.pack(side='left', expand=True, padx=(0, 0), fill="x")
    Pull_log_frame.pack(expand=True, fill="both", pady=5)

    # 运行主循环
    root.mainloop()
