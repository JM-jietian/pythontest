import logging
import os
import re
import subprocess
import threading
import json
from django.views.generic import TemplateView
from django.shortcuts import render, redirect
from django.urls import reverse
from openpyxl import Workbook

# Create your views here.
def connection_status(status_func):
    # 定义装饰器函数，判断设备连接状态
    def device(*args, **kwargs):
        devices = [line.split('\t')[0].strip() for line in os.popen('adb devices').readlines()[1:] if
                   line.split('\t')[0].strip()]
        num_devices = len(devices)
        if num_devices == 0:
            print('暂无设备连接')
        elif num_devices > 1:
            print('当前连接多台设备，仅支持单设备操作，请检查！')
        else:
            status_func(*args, **kwargs)

    return device

class InitiateView(TemplateView):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.log_file = 'LOG.log'
        # 初始化日志记录器
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        # 创建文件处理器，并设置级别为INFO
        fh = logging.FileHandler(self.log_file, mode='a', encoding='utf-8')  # 存储到日志文件，a追加
        # fh = logging.StreamHandler()  # 控制台输出
        fh.setLevel(logging.INFO)
        # 创建并设置日志格式
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        # 添加处理器到日志记录器
        self.logger.addHandler(fh)

        self.start_frame = None  # 初始帧
        self.end_frame = None  # 结束帧
        self.paths = {}  # 文件路径
        self.image_files = []  # 存储图片文件
        self.screen_recording_time = None  # 录屏时长
        self.startup_duration = None # 耗时
        self.len_image_files = None # 图片总数


    def sub(self, cmd, callback=None):
        # 封装处理subprocess
        self.logger.info(f'执行指令：{cmd}')
        try:
            output = subprocess.check_output(cmd, shell=True, stderr=subprocess.STDOUT).decode('utf-8', errors='ignore')
            self.logger.info(output)
            if callback:
                callback(output)
            return output.strip()
        except subprocess.CalledProcessError as e:
            self.logger.error("Command '%s' failed with error: %s", cmd, e.output.decode('utf-8', errors='ignore'))
            if callback:
                callback(None)

    @staticmethod
    def find_file_partial(file_pattern, search_path='/'):
        # 输出指定文件完整路径
        found = None
        found_event = threading.Event()

        def search_files():
            nonlocal found
            for root, dirs, files in os.walk(search_path):
                for file_name in files:
                    if re.search(file_pattern, file_name):
                        found = os.path.abspath(root)  # 返回文件的完整路径
                        found_event.set()  # 通知主线程
                        return

        thread = threading.Thread(target=search_files)
        thread.start()
        found_event.wait()  # 等待搜索完成
        thread.join()  # 确保线程结束
        return found

    @staticmethod
    def write_to_excel(data, file_path):
        # 创建一个工作簿
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        # 写入标题行
        for col, value in enumerate(data[0]):
            ws.cell(row=1, column=col + 1, value=value)
        # 写入数据行
        for row, row_data in enumerate(data[1:], start=2):  # 从第二行开始写入
            for col, value in enumerate(row_data):
                ws.cell(row=row, column=col + 1, value=value)

        # 保存工作簿到文件
        wb.save(file_path)

    def delete_files_with_extension(self, folder_path, extension):
        # 检查文件夹路径是否存在
        if not os.path.exists(folder_path):
            self.logger.error("文件夹路径不存在")
            return

        # 确保扩展名以点开头
        if not extension.startswith('.'):
            extension = '.' + extension

        # 遍历文件夹及其子文件夹中的所有文件
        for root, dirs, files in os.walk(folder_path):
            for filename in files:
                file_path = os.path.join(root, filename)

                # 检查文件后缀名
                if filename.endswith(extension):
                    try:
                        os.unlink(file_path)
                        self.logger.info(f"已删除文件：{file_path}")
                    except Exception as e:
                        self.logger.error(f"删除文件 {file_path} 时出错：{e}")

    def add_to_path_if_not_exists(self, directory_path):
        # 获取系统PATH环境变量
        paths = os.environ.get('PATH', '')
        # 根据操作系统使用正确的路径分隔符
        path_separator = os.pathsep
        # 检查目录路径是否已经在PATH中
        if directory_path in paths.split(path_separator):
            self.logger.info(f"路径 {directory_path} 已在PATH中。")
            return

        # 如果路径不在PATH中，将其添加进去
        if not paths.endswith(path_separator):
            new_path = paths + path_separator + directory_path
        else:
            new_path = paths + directory_path

        # 更新环境变量，临时添加，只会影响当前进程及其子进程
        os.environ['PATH'] = new_path
        self.logger.info(f"路径 {directory_path} 临时添加到PATH环境变量中。")

    @connection_status
    def initiate(self):
        self.logger.info('正在启动投屏 ······\n注意：操作完成后关闭投屏即可')
        self.paths.update(
            {
                "scrcpy": self.find_file_partial("scrcpy-console.bat"),
                "ffmpeg": self.find_file_partial("ffmpeg_initiate"),
                "video": self.find_file_partial("ffmpeg_video"),
                "image": self.find_file_partial("ffmpeg_Image")
            }
        )
        # 判断 scrcpy，ffmpeg是否添加了环境变量，如果没有则临时添加
        self.add_to_path_if_not_exists(self.paths['scrcpy'])
        self.add_to_path_if_not_exists(self.paths['ffmpeg'])
        # 开始录屏并保存视频
        self.sub(rf"scrcpy --record {self.paths['video']}\video.mp4")
        # 获取录屏时长
        self.screen_recording_time = self.sub(
            rf'ffprobe -v quiet -show_entries format=duration -of default=noprint_wrappers=1:nokey=1 "{self.paths["video"]}\video.mp4"')
        self.logger.info("录屏时长: %s 秒" % self.screen_recording_time)
        # 拆解录屏，生成每一帧的图片
        self.delete_files_with_extension(self.paths['image'], '.png')  # 拆解前先删除之前的图片文件
        self.sub(rf"ffmpeg -i {self.paths['video']}\video.mp4 -r 10 {self.paths['image']}\%d.png")
        # 获取拆解后图片总数
        self.image_files = [os.path.join(self.paths['image'], f) for f in os.listdir(self.paths['image'])
                            if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif'))]
        self.logger.info("拆解图片： %s 张" % len(self.image_files))

    @connection_status
    def compute(self):
        # 计算启动耗时
        self.logger.info('开始计算启动耗时，请稍后！')
        self.start_frame += 1
        self.end_frame += 1
        if self.screen_recording_time and self.len_image_files and self.start_frame and self.end_frame:
            # 计算时间：视频总时间/拆解视频图片总帧数*间隔帧数
            self.startup_duration = float(self.screen_recording_time) / int(self.len_image_files) * int(
                self.end_frame - self.start_frame)
            self.logger.info("耗时：%s 秒" % self.startup_duration)
        else:
            self.logger.error("参数异常！\n\n录屏时长: %s \n图片总数: %s \n开始帧: %s \n结束帧: %s \n"
                  % (self.screen_recording_time, self.len_image_files, self.start_frame, self.end_frame))

    def get_context_data(self, **kwargs):
        # 调用基类的 get_context_data 方法
        context = super().get_context_data(**kwargs)
        devices = [line.split('\t')[0].strip() for line in os.popen('adb devices').readlines()[1:] if line.split('\t')[0].strip()]
        context['devices'] = devices
        return context

    def get(self, request, *args, **kwargs):
        # 处理get请求

        context = self.get_context_data(**kwargs)
        return render(request, 'index.html',context)

    def post(self, request, *args, **kwargs):
        # 处理post请求

        # 查看get请求发送的表单信息
        device_id = request.POST.get("device_id")
        # 查看ajax发送的表单信息
        start_frame = request.POST.get('startFrame')
        end_frame = request.POST.get('endFrame')

        if device_id:
            self.initiate()

            # 图片路径拆分
            image_files = []
            for file in self.image_files:
                image_file = list(file.split('\\')[-2:])
                image_files.append(image_file)
            # 转换为JSON格式的字符串
            image_files_json = json.dumps(image_files, indent=4)  # indent参数用于美化输出，使JSON更易读

            # 更新上下文数据
            context = self.get_context_data(**kwargs)
            context["screen_recording_time"] = self.screen_recording_time,  # 录屏时长
            context["len_image_files"] = len(self.image_files),  # 拆解后图片总数
            context["image_files_json"] = image_files_json  # 图片路径

            return render(request, 'image.html', context)

        elif start_frame and end_frame:

            self.start_frame = int(start_frame)  # 初始帧
            self.end_frame = int(end_frame)  # 结束帧

            self.screen_recording_time = request.POST.get("screen_recording_time")  # 录屏时长
            self.len_image_files = request.POST.get("len_image_files")  # 图片总数

            self.compute()

            # 更新上下文数据
            context = self.get_context_data(**kwargs)
            context["startup_duration"] = str(self.startup_duration)  # 耗时

            return render(request, 'index.html', context)
        else:
            return redirect(reverse('index'))